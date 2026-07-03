import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Optional


class ExcelProcessor:

    def read_excel(self, file_name: str) -> Optional[List[List[object]]]:
        data: List[List[object]] = []
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.data_type == 'n':
                        row_data.append(int(cell.value))
                    elif cell.data_type == 's':
                        row_data.append(cell.value)
                    else:
                        row_data.append(None)
                data.append(row_data)
            return data
        except Exception:
            return None

    def write_excel(self, data: List[List[object]], file_name: str) -> bool:
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row_idx, row in enumerate(data, 1):
                sheet.cell(row=row_idx, column=1, value=row[0])
                sheet.cell(row=row_idx, column=2, value=row[1])
            # Extend columns if necessary
            max_columns = max(len(row) for row in data)
            for col in range(1, max_columns + 1):
                sheet.cell(row=1, column=col, value=f"Column {col}")
            workbook.save(file_name)
            return True
        except Exception:
            return False

    def process_excel_data(self, N: int, save_file_name: str) -> Optional[str]:
        data = self.read_excel(save_file_name)
        if data is None or N >= len(data[0]) if data else True:
            return None
        new_data = []
        for row in data:
            new_row = row.copy()
            value = row[N]
            if isinstance(value, str):
                new_row.append(value.upper())
            else:
                new_row.append(value)
            new_data.append(new_row)
        new_file_name = save_file_name.replace('.xlsx', '_process.xlsx')
        if self.write_excel(new_data, new_file_name):
            return new_file_name
        return None