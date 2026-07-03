import openpyxl

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name):
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb["Sheet1"]
            data = []
            max_row = sheet.max_row
            max_col = sheet.max_column
            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if isinstance(value, int):
                        row_data.append(value)
                    elif isinstance(value, str):
                        row_data.append(value)
                data.append(row_data)
            wb.close()
            return data
        except:
            return []

    def write_excel(self, data, file_name):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for row_idx, row in enumerate(data):
                for col_idx, value in enumerate(row):
                    if isinstance(value, int):
                        sheet.cell(row=row_idx + 1, column=col_idx + 1).value = value
                    elif isinstance(value, str):
                        sheet.cell(row=row_idx + 1, column=col_idx + 1).value = value
            wb.save(file_name)
            wb.close()
            return 1
        except:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")
        
        new_data = []
        for row in data:
            new_row = row.copy()
            element = row[N]
            if type(element) is str:
                if not element[0].isdigit():
                    new_str = ''.join(chr(ord(c) - 32) if ord(c) > 90 else c for c in element)
                    new_row.append(new_str)
                else:
                    new_row.append(element[0])
            else:
                new_row.append(element)
            new_data.append(new_row)
        
        if '.' in save_file_name:
            base = save_file_name[:save_file_name.rfind('.')]
        else:
            base = save_file_name
        new_file_name = base + "_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)