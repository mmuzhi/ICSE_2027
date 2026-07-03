import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelProcessor:

    def __init__(self):
        pass

    def readExcel(self, fileName):
        try:
            wb = load_workbook(fileName, read_only=True)
            sheet = wb.worksheets[0]
            data = []
            for row in sheet.rows:
                rowData = []
                for cell in row:
                    if cell.data_type == 's':
                        rowData.append(cell.value)
                    elif cell.data_type == 'n':
                        try:
                            rowData.append(int(cell.value))
                        except:
                            rowData.append(None)
                    else:
                        rowData.append(None)
                data.append(rowData)
            wb.close()
            return data
        except Exception as e:
            return None

    def writeExcel(self, data, fileName):
        try:
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for i, rowData in enumerate(data):
                for j, value in enumerate(rowData):
                    cell = sheet.cell(row=i+1, column=j+1)
                    if isinstance(value, str):
                        cell.value = value
                    elif isinstance(value, int):
                        cell.value = value
            wb.save(fileName)
            return True
        except Exception as e:
            return False

    def processExcelData(self, N, saveFileName):
        data = self.readExcel(saveFileName)
        if data is None or not data or N >= len(data[0]):
            return None
        newData = []
        for row in data:
            newRow = list(row)
            value = newRow[N]
            if isinstance(value, str):
                newRow.append(value.upper())
            else:
                newRow.append(value)
            newData.append(newRow)
        newFileName = saveFileName.replace(".xlsx", "_process.xlsx")
        success = self.writeExcel(newData, newFileName)
        return newFileName if success else None