import openpyxl
import os

class ExcelProcessor:
    def read_excel(self, file_name):
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
                row_data = []
                for cell_value in row:
                    if cell_value is None:
                        continue
                    if isinstance(cell_value, str):
                        if cell_value.lstrip('-').isdigit() and cell_value != '-' and cell_value.lstrip('-').strip().isdigit():
                            try:
                                row_data.append(int(cell_value))
                            except ValueError:
                                row_data.append(cell_value)
                        else:
                            row_data.append(cell_value)
                    else:
                        row_data.append(cell_value)
                data.append(row_data)
            return data
        except Exception as e:
            return []

    def write_excel(self, data, file_name):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for row_idx, row in enumerate(data):
                for col_idx, cell_value in enumerate(row):
                    cell = sheet.cell(row=row_idx+1, column=col_idx+1)
                    if isinstance(cell_value, int):
                        cell.value = cell_value
                    elif isinstance(cell_value, str):
                        cell.value = cell_value
            wb.save(file_name)
            return 1
        except Exception as e:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or not data[0] or N >= len(data[0]):
            return (0, "")
        
        new_data = []
        for row in data:
            new_row = list(row)
            col_value = row[N]
            if isinstance(col_value, str) and col_value:
                if col_value[0].isdigit():
                    new_row.append(col_value[0])
                else:
                    s = col_value
                    for i in range(len(s)):
                        if s[i] > 'Z':
                            s_list = list(s)
                            if ord(s[i]) > 90:
                                s_list[i] = chr(ord(s[i]) - 32)
                            s = ''.join(s_list)
                    new_row.append(s)
            else:
                new_row.append(col_value)
            new_data.append(new_row)
        
        base = save_file_name[:save_file_name.rfind('.')] if '.' in save_file_name else save_file_name
        new_file_name = base + "_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)