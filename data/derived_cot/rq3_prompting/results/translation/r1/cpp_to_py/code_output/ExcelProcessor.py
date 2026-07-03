import openpyxl
import os
from typing import List, Tuple, Union

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name: str) -> List[List[Union[str, int]]]:
        data = []
        try:
            wb = openpyxl.load_workbook(file_name, data_only=True)
            sheet = wb["Sheet1"]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
                row_data = []
                for cell in row:
                    v = cell.value
                    if isinstance(v, int):
                        row_data.append(v)
                    elif isinstance(v, str):
                        row_data.append(v)
                data.append(row_data)
            wb.close()
        except Exception:
            return []
        return data

    def write_excel(self, data: List[List[Union[str, int]]], file_name: str) -> int:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for i, row_data in enumerate(data, start=1):
                for j, val in enumerate(row_data, start=1):
                    cell = sheet.cell(row=i, column=j)
                    cell.value = val
            wb.save(file_name)
            wb.close()
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N: int, save_file_name: str) -> Tuple[int, str]:
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")

        new_data = []
        for row in data:
            new_row = list(row)
            val = row[N]
            if isinstance(val, str):
                if not val[0].isdigit():
                    new_row.append(val.upper())
                else:
                    new_row.append(val[0])
            else:
                new_row.append(val)
            new_data.append(new_row)

        base, ext = os.path.splitext(save_file_name)
        new_file_name = base + "_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)