import openpyxl


class ExcelProcessor:

    def __init__(self):
        pass

    def read_excel(self, filename: str):
        """Returns list of rows, each row is list of values (str, int, or None)."""
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.worksheets[0]
            data = []
            for row in sheet.iter_rows(values_only=False):
                row_data = []
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        row_data.append(val)
                    elif isinstance(val, (int, float)):
                        # Match Java's (int) cast – truncate toward zero
                        row_data.append(int(val))
                    else:
                        row_data.append(None)
                data.append(row_data)
            return data
        except Exception:
            return None

    def write_excel(self, data, filename: str) -> bool:
        """Writes data (list of lists of str/int) to .xlsx file. Returns True on success."""
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for i, row_data in enumerate(data):
                row = sheet.cell(row=i + 1, column=1)  # start at row 1
                for j, value in enumerate(row_data):
                    cell = sheet.cell(row=i + 1, column=j + 1)
                    if isinstance(value, str):
                        cell.value = value
                    elif isinstance(value, int):
                        cell.value = value
                    # other types ignored, same as Java (no else)
            wb.save(filename)
            return True
        except Exception:
            return False

    def process_excel_data(self, n: int, save_filename: str):
        """
        Reads Excel file, appends a column that is the uppercase of column n (if string) or the original value,
        then writes to a new file with suffix '_process.xlsx'.
        Returns new filename on success, None otherwise.
        """
        data = self.read_excel(save_filename)
        if data is None or n >= len(data[0]):
            return None

        new_data = []
        for row in data:
            new_row = list(row)          # copy
            value = row[n]
            if isinstance(value, str):
                new_row.append(value.upper())
            else:
                new_row.append(value)
            new_data.append(new_row)

        new_filename = save_filename.replace(".xlsx", "_process.xlsx")
        success = self.write_excel(new_data, new_filename)
        return new_filename if success else None