import openpyxl
from typing import List, Any, Optional

def read_excel(file_name: str) -> Optional[List[List[Any]]]:
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.worksheets[0]
        data = []
        for row in sheet:
            row_data = []
            for cell in row:
                if cell.data_type == 's':
                    row_data.append(cell.value)
                elif cell.data_type == 'n':
                    row_data.append(int(cell.value))
                else:
                    row_data.append(None)
            data.append(row_data)
        return data
    except FileNotFoundError:
        return None
    except Exception as e:
        return None

def write_excel(data: List[List[Any]], file_name: str) -> bool:
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                cell = sheet.cell(row=i+1, column=j+1)
                if isinstance(value, str):
                    cell.value = value
                elif isinstance(value, int):
                    cell.value = value
        workbook.save(file_name)
        return True
    except Exception:
        return False

def process_excel_data(N: int, save_filename: str) -> Optional[str]:
    data = read_excel(save_filename)
    if data is None:
        return None
    try:
        if N >= len(data[0]):
            return None
    except IndexError:
        return None
    new_data = []
    for row in data:
        new_row = row.copy()
        value = row[N]
        if isinstance(value, str):
            new_row.append(value.upper())
        else:
            new_row.append(value)
        new_data.append(new_row)
    new_file_name = save_filename.replace(".xlsx", "_process.xlsx")
    success = write_excel(new_data, new_file_name)
    return new_file_name if success else None