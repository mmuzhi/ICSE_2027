import openpyxl
from typing import List, Union, Tuple
import os

class ExcelProcessor:
    def read_excel(self, file_name: str) -> List[List[Union[str, int]]]:
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb['Sheet1']
            data = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                       min_col=1, max_col=sheet.max_column, 
                                       values_only=True):
                row_data = []
                for cell in row:
                    if cell is None:
                        continue
                    if isinstance(cell, int):
                        row_data.append(cell)
                    elif isinstance(cell, str):
                        row_data.append(cell)
                    else:
                        row_data.append(str(cell))
                if row_data:
                    data.append(row_data)
            wb.close()
            return data
        except Exception:
            return []

    def write_excel(self, data: List[List[Union[str, int]]], file_name: str) -> int:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Sheet1'
            for row_idx, row in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if isinstance(cell_value, int):
                        sheet.cell(row=row_idx, column=col_idx, value=str(cell_value))
                    else:
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            wb.save(file_name)
            wb.close()
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N: int, save_file_name: str) -> Tuple[int, str]:
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")
        
        new_data = []
        for row in data:
            new_row = row.copy()
            if N < len(row):
                cell_value = row[N]
                if isinstance(cell_value, str):
                    if not cell_value[0].isdigit():
                        new_str = list(cell_value)
                        for i in range(len(new_str)):
                            if new_str[i] > 'Z':
                                new_str[i] = chr(ord(new_str[i]) - 32)
                        new_str = ''.join(new_str)
                        new_row.append(new_str)
                    else:
                        new_row.append(cell_value[:1])
                else:
                    new_row.append(cell_value)
            new_data.append(new_row)
        
        base_name = os.path.splitext(save_file_name)[0]
        new_file_name = f"{base_name}_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)