import openpyxl


class ExcelProcessor:
    def read_excel(self, file_name):
        data = []
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb["Sheet1"]
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                       min_col=1, max_col=sheet.max_column):
                row_data = []
                for cell in row:
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, int) and not isinstance(cell.value, bool):
                        row_data.append(int(cell.value))
                    elif isinstance(cell.value, str):
                        row_data.append(cell.value)
                data.append(row_data)
            wb.close()
        except Exception:
            return []
        return data

    def write_excel(self, data, file_name):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for row_index, row in enumerate(data):
                for col_index, value in enumerate(row):
                    cell = sheet.cell(row=row_index + 1, column=col_index + 1)
                    if isinstance(value, int):
                        cell.value = value
                    else:
                        cell.value = value
            wb.save(file_name)
            wb.close()
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")

        new_data = []
        for row in data:
            new_row = list(row)
            val = row[N]
            if isinstance(val, str):
                if not ('0' <= val[0] <= '9'):
                    now = list(val)
                    for i in range(len(now)):
                        if now[i] > 'Z':
                            now[i] = chr(ord(now[i]) - 32)
                    new_row.append(''.join(now))
                else:
                    new_row.append(val[:1])
            else:
                new_row.append(val)
            new_data.append(new_row)

        dot_pos = save_file_name.rfind('.')
        new_file_name = save_file_name[:dot_pos] + "_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)