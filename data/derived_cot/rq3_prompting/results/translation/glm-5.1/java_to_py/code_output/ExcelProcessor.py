import openpyxl


class ExcelProcessor:

    def __init__(self):
        pass

    def readExcel(self, fileName):
        data = []
        try:
            workbook = openpyxl.load_workbook(fileName, data_only=True)
            sheet = workbook.worksheets[0]
            for row in sheet.iter_rows():
                rowData = []
                for cell in row:
                    if cell.value is None:
                        rowData.append(None)
                    elif isinstance(cell.value, str):
                        rowData.append(cell.value)
                    elif isinstance(cell.value, (int, float)):
                        rowData.append(int(cell.value))
                    else:
                        rowData.append(None)
                data.append(rowData)
            workbook.close()
        except (IOError, OSError, Exception):
            return None
        return data

    def writeExcel(self, data, fileName):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            for i, rowData in enumerate(data):
                for j, value in enumerate(rowData):
                    cell = sheet.cell(row=i + 1, column=j + 1)
                    if isinstance(value, str):
                        cell.value = value
                    elif isinstance(value, int):
                        cell.value = value
            workbook.save(fileName)
            workbook.close()
            return True
        except (IOError, OSError, Exception):
            return False

    def processExcelData(self, N, saveFileName):
        data = self.readExcel(saveFileName)
        if data is None or N >= len(data[0]):
            return None
        newData = []
        for row in data:
            newRow = list(row)
            value = row[N]
            if isinstance(value, str):
                newRow.append(value.upper())
            else:
                newRow.append(value)
            newData.append(newRow)
        newFileName = saveFileName.replace(".xlsx", "_process.xlsx", 1)
        success = self.writeExcel(newData, newFileName)
        return newFileName if success else None