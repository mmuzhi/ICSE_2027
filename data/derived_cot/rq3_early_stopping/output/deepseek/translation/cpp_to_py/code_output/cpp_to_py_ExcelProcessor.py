import openpyxl
import os
from typing import List, Union, Tuple

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name: str) -> List[List[Union[str, int]]]:
        data = []
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb["Sheet1"]
            for row in sheet.iter_rows(min_row=1, values_only=False):
                row_data = []
                for cell in row:
                    if cell.value is None:
                        continue
                    if isinstance(cell.value, int):
                        row_data.append(cell.value)
                    elif isinstance(cell.value, str):
                        row_data.append(cell.value)
                    elif isinstance(cell.value, float):
                        if cell.value == int(cell.value):
                            row_data.append(int(cell.value))
                        else:
                            continue
                    else:
                        continue
                if row_data:
                    data.append(row_data)
            wb.close()
        except Exception:
            return []
        return data

    def write_excel(self, data: List[List[Union[str, int]]], file_name: str) -> int:
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for row_idx, row in enumerate(data, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
            wb.save(file_name)
            wb.close()
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N: int, save_file_name: str) -> Tuple[int, str]:
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")

        new_data = []
        for row in data:
            new_row = list(row)
            value = row[N]
            if isinstance(value, str):
                if not value[0].isdigit():
                    now = value
                    now_list = list(now)
                    for i in range(len(now_list)):
                        if now_list[i] > 'Z':
                            now_list[i] = chr(ord(now_list[i]) - 32)
                    new_row.append(''.join(now_list))
                else:
                    new_row.append(value[0])
            else:
                new_row.append(value)
            new_data.append(new_row)

        base, ext = os.path.splitext(save_file_name)
        new_file_name = base + "_process.xlsx"
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)