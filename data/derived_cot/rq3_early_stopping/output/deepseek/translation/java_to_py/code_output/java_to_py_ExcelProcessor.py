import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name: str):
        """Reads the first sheet of an .xlsx file and returns a list of lists of objects.
        Returns None on failure."""
        try:
            workbook = openpyxl.load_workbook(file_name, data_only=True)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_type = cell.data_type
                    if cell_type == 's':  # string
                        row_data.append(cell.value)
                    elif cell_type == 'n':  # numeric
                        row_data.append(int(cell.value))
                    else:
                        row_data.append(None)
                data.append(row_data)
            workbook.close()
            return data
        except Exception:
            return None

    def write_excel(self, data, file_name: str):
        """Writes the data to a new .xlsx file.
        Returns True on success, False on failure."""
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            for i, row_data in enumerate(data, start=1):
                row = sheet.cell(row=i, column=1)  # just to start a row, we'll set columns
                for j, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=i, column=j)
                    if isinstance(value, str):
                        cell.value = value
                    elif isinstance(value, int):
                        cell.value = value
            workbook.save(file_name)
            workbook.close()
            return True
        except Exception:
            return False

    def process_excel_data(self, n: int, save_file_name: str):
        """Reads an Excel file, processes column N, and writes a new file.
        Returns the new file name on success, or None on failure."""
        data = self.read_excel(save_file_name)
        if data is None or n >= len(data[0]):
            return None
        new_data = []
        for row in data:
            new_row = list(row)  # copy
            value = row[n]
            if isinstance(value, str):
                new_row.append(value.upper())
            else:
                new_row.append(value)
            new_data.append(new_row)
        new_file_name = save_file_name.replace(".xlsx", "_process.xlsx")
        success = self.write_excel(new_data, new_file_name)
        return new_file_name if success else None