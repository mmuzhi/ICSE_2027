import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Optional

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name: str) -> Optional[List[List[object]]]:
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.data_type == 'n':
                        # Numeric cell
                        row_data.append(cell.value)
                    elif cell.data_type == 's':
                        # String cell
                        row_data.append(cell.value)
                    else:
                        row_data.append(None)
                data.append(row_data)
            return data
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

    def write_excel(self, data: List[List[object]], file_name: str) -> bool:
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    if i < len(data) and j < len(row_data):
                        cell = sheet.cell(row=i+1, column=j+1)
                        if isinstance(value, str):
                            cell.value = value
                        elif isinstance(value, int):
                            cell.value = value
                        else:
                            cell.value = None
            workbook.save(file_name)
            return True
        except Exception as e:
            print(f"Error writing Excel file: {e}")
            return False

    def process_excel_data(self, N: int, save_file_name: str) -> Optional[str]:
        # Read the Excel file
        data = self.read_excel(save_file_name)
        if data is None:
            return None
        
        # Check if N is within the bounds of the first row
        if N >= len(data[0]):
            return None
        
        # Process each row: convert the N-th column to uppercase if it's a string, else leave as is
        new_data = []
        for row in data:
            new_row = row.copy()  # Create a copy of the row
            # Get the value at column N
            value = row[N]
            if isinstance(value, str):
                new_row.append(value.upper())
            else:
                new_row.append(value)
            new_data.append(new_row)
        
        # Create a new file name by replacing the extension
        new_file_name = save_file_name.replace('.xlsx', '_process.xlsx')
        success = self.write_excel(new_data, new_file_name)
        return new_file_name if success else None