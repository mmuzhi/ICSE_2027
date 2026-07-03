from openpyxl import load_workbook
from openpyxl import Workbook
import re

class ExcelProcessor:
    def read_excel(self, file_name):
        try:
            wb = load_workbook(filename=file_name)
            sheet = wb.active
            
            data = []
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                row_data = []
                for cell in row:
                    if cell.data_type == 'n':
                        if cell.value is not None and isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value.is_integer()):
                            row_data.append(int(cell.value))
                        else:
                            raise ValueError(f"Non-integer number found: {cell.value}")
                    elif cell.data_type == 's':
                        row_data.append(cell.value)
                    else:
                        raise ValueError(f"Unexpected cell type: {cell.data_type}")
                data.append(row_data)
            return data
        except Exception as e:
            return []

    def write_excel(self, data, file_name):
        try:
            wb = Workbook()
            sheet = wb.active
            
            for row_idx, row in enumerate(data, start=1):
                for col_idx, cell_value in enumerate(row, start=1):
                    if isinstance(cell_value, int):
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    elif isinstance(cell_value, str):
                        sheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    else:
                        raise ValueError(f"Unexpected cell type: {type(cell_value)}")
            
            wb.save(file_name)
            return 1
        except Exception as e:
            return 0

    def process_excel_data(self, N, save_file_name):
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")
        
        new_data = []
        for row in data:
            if N >= len(row):
                new_row = row[:]
            else:
                new_row = row[:]
                cell_value = new_row[N]
                if isinstance(cell_value, str):
                    if cell_value and cell_value[0].isdigit():
                        new_row[N] = cell_value[0]
                    else:
                        new_str = ''.join([char if not (ord(char) >= ord('a') and ord(char) <= ord('Z')) else chr(ord(char)-32) for char in cell_value])
                        new_row[N] = new_str
                elif isinstance(cell_value, int):
                    pass
                else:
                    pass
            new_data.append(new_row)
        
        new_file_name = re.sub(r'\.xlsx$', '_process.xlsx', save_file_name)
        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)