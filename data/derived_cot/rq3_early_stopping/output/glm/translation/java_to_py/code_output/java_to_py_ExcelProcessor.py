import openpyxl

class ExcelProcessor:

    def __init__(self):
        pass

    def readExcel(self, fileName):
        data = []
        try:
            wb = openpyxl.load_workbook(fileName)
            sheet = wb.worksheets[0]
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.value is None:
                        row_data.append(None)
                    elif cell.data_type == 's':
                        row_data.append(cell.value)
                    elif cell.data_type == 'n':
                        row_data.append(int(cell.value))
                    else:
                        row_data.append(None)
                data.append(row_data)
        except OSError:
            return None
        return data

    def writeExcel(self, data, fileName):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Sheet1"
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    if isinstance(value, str):
                        sheet.cell(row=i + 1, column=j + 1, value=value)
                    elif isinstance(value, int):
                        sheet.cell(row=i + 1, column=j + 1, value=value)
            wb.save(fileName)
            return True
        except OSError:
            return False

    def processExcelData(self, N, saveFileName):
        data = self.readExcel(saveFileName)
        if data is None or N >= len(data[0]):
            return None
        new_data = []
        for row in data:
            new_row = list(row)
            value = row[N]
            if isinstance(value, str):
                new_row.append(value.upper())
            else:
                new_row.append(value)
            new_data.append(new_row)
        newFileName = saveFileName.replace(".xlsx", "_process.xlsx")
        success = self.writeExcel(new_data, newFileName)
        return newFileName if success else None