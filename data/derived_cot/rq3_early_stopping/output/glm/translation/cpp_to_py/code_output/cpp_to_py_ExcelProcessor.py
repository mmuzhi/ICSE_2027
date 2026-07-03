import openpyxl
from typing import List, Union, Tuple

class ExcelProcessor:
    def __init__(self):
        pass

    def read_excel(self, file_name: str) -> List[List[Union[str, int]]]:
        data = []
        try:
            doc = openpyxl.load_workbook(file_name)
            sheet = doc["Sheet1"]

            for row_idx in range(1, sheet.max_row + 1):
                row_data = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if type(val) is int:
                        row_data.append(val)
                    elif type(val) is str:
                        row_data.append(val)
                data.append(row_data)

            doc.close()
        except Exception:
            return []
        return data

    def write_excel(self, data: List[List[Union[str, int]]], file_name: str) -> int:
        try:
            doc = openpyxl.Workbook()
            sheet = doc.active
            sheet.title = "Sheet1"

            for row_idx, row in enumerate(data):
                for col_idx, val in enumerate(row):
                    cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                    if type(val) is int:
                        cell.value = val
                    else:
                        cell.value = str(val)

            doc.save(file_name)
            doc.close()
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N: int, save_file_name: str) -> Tuple[int, str]:
        data = self.read_excel(save_file_name)
        if not data or N >= len(data[0]):
            return (0, "")

        new_data = []
        for row in data:
            new_row = list(row)
            val = row[N]

            if type(val) is str:
                if val and not val[0].isdigit():
                    now = list(val)
                    for i in range(len(now)):
                        if ord(now[i]) > ord('Z'):
                            now[i] = chr(ord(now[i]) - 32)
                    new_row.append("".join(now))
                else:
                    new_row.append(val[:1])
            else:
                new_row.append(val)

            new_data.append(new_row)

        dot_idx = save_file_name.rfind('.')
        if dot_idx != -1:
            base_name = save_file_name[:dot_idx]
        else:
            base_name = save_file_name

        new_file_name = base_name + "_process.xlsx"

        success = self.write_excel(new_data, new_file_name)
        return (success, new_file_name)